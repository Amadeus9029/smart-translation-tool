import os
import time
import pandas as pd
from openpyxl import load_workbook
import logging
import concurrent.futures
from tqdm import tqdm
from openai import OpenAI
import threading
import shutil

# 在文件开头添加 SUPPORTED_LANGUAGES 定义
SUPPORTED_LANGUAGES = {
    "中文": "Chinese",
    "英语": "English",
    "法语": "French",
    "德语": "German",
    "西班牙语": "Spanish",
    "意大利语": "Italian",
    "日语": "Japanese",
    "韩语": "Korean",
    "俄语": "Russian",
    "葡萄牙语": "Portuguese",
    "阿拉伯语": "Arabic",
    "荷兰语": "Dutch",
    "波兰语": "Polish",
    "土耳其语": "Turkish",
    "瑞典语": "Swedish",
    "丹麦语": "Danish",
    "芬兰语": "Finnish",
    "希腊语": "Greek",
    "捷克语": "Czech",
    "匈牙利语": "Hungarian",
    "罗马尼亚语": "Romanian",
    "保加利亚语": "Bulgarian",
    "印尼语": "Indonesian",
    "泰语": "Thai",
    "越南语": "Vietnamese"
}

# 修改日志配置，只输出到控制台
logger = logging.getLogger()
console_handler = logging.StreamHandler()
console_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
logger.addHandler(console_handler)
logger.setLevel(logging.INFO)

# DeepSeek API配置
DEEPSEEK_BASE_URL = "https://api.deepseek.com"

# 默认配置参数
DEFAULT_MAX_WORKERS = 5      # 默认并发线程数
DEFAULT_BATCH_SIZE = 10      # 默认每批处理的条目数
DEFAULT_MAX_RETRIES = 3      # 默认最大重试次数
DEFAULT_SAVE_INTERVAL = 100  # 默认每处理100个单元格保存一次
DEFAULT_PROGRESS_INTERVAL = 10  # 默认每翻译10个单元格显示一次进度

# 全局配置变量
max_workers = DEFAULT_MAX_WORKERS
batch_size = DEFAULT_BATCH_SIZE
max_retries = DEFAULT_MAX_RETRIES
save_interval = DEFAULT_SAVE_INTERVAL
progress_interval = DEFAULT_PROGRESS_INTERVAL

# 创建锁对象用于线程安全操作
excel_lock = threading.Lock()
progress_lock = threading.Lock()
translated_count = 0  # 全局计数器，记录已翻译的单元格数
last_progress_report = 0  # 上次显示进度的计数

# 添加进度回调函数
progress_callback = None

# 全局变量
translation_cancelled = False
api_key = None  # 存储API Key的全局变量
total_tasks = 0  # 添加全局变量

def set_translation_cancelled(value):
    """设置翻译取消状态"""
    global translation_cancelled
    translation_cancelled = value

def translate_batch_with_reference(batch_data, target_lang, reference_lang, retry_count=0):
    """带参考翻译的批量翻译"""
    if not batch_data:
        return []
    
    try:
        if not api_key:
            raise ValueError("API Key未设置")
            
        client = OpenAI(api_key=api_key, base_url=DEEPSEEK_BASE_URL)
        
        # 构建批量翻译请求
        batch_prompts = []
        for i, (source_text, ref_text) in enumerate(batch_data):
            if ref_text:  # 如果有参考文本
                batch_prompts.append(f"{i+1}. 原文: {source_text}\n   {reference_lang}参考翻译: {ref_text}")
            else:  # 如果没有参考文本
                batch_prompts.append(f"{i+1}. 原文: {source_text}")
        
        batch_text = "\n\n".join(batch_prompts)
        
        prompt = f"""请将以下{len(batch_data)}条文本翻译成{target_lang}。
如果提供了参考翻译，请确保翻译的内容与参考翻译在语义上保持一致。
仅返回翻译结果，格式为"1. [翻译结果1]"，"2. [翻译结果2]"等，不要有额外解释。

{batch_text}"""

        if translation_cancelled:
            return [("[已取消]", "[已取消]") for _ in batch_data]

        response = client.chat.completions.create(
            model="deepseek-chat",
            messages=[
                {"role": "system", "content": "你是一个精通多领域的翻译专家。只翻译文本，不添加任何解释或附加文本。返回结果保持简洁。"},
                {"role": "user", "content": prompt}
            ],
            temperature=0.3,
            max_tokens=2000,
        )
        
        # 解析返回结果
        result = response.choices[0].message.content.strip()
        translated_texts = []
        
        # 处理返回的格式，试图匹配"数字. 翻译内容"格式
        lines = result.split("\n")
        for i, (source_text, _) in enumerate(batch_data):
            found = False
            for line in lines:
                line = line.strip()
                if line.startswith(f"{i+1}. ") or line.startswith(f"{i+1}."):
                    translated = line[line.find(".")+1:].strip()
                    translated_texts.append((source_text, translated))
                    found = True
                    break
            
            if not found:
                translated_texts.append((source_text, "[格式错误]"))
        
        return translated_texts
        
    except Exception as e:
        logger.error(f"批量翻译出错: {e}")
        if retry_count < max_retries:
            logger.info(f"第{retry_count+1}次重试批量翻译...")
            time.sleep(1)
            return translate_batch_with_reference(batch_data, target_lang, reference_lang, retry_count + 1)
        # 失败时返回错误信息
        return [(text[0] if isinstance(text, tuple) else text, "[翻译错误]") for text, _ in batch_data]

def update_progress_status(current, total, finished=False):
    """更新进度状态"""
    global last_progress_report
    
    with progress_lock:
        if current - last_progress_report >= progress_interval or finished:
            last_progress_report = current
            # 使用回调更新进度
            if progress_callback:
                try:
                    progress_callback(current, total, finished)
                except Exception as e:
                    logger.error(f"更新进度时出错: {e}")

def count_already_translated(ws, valid_rows, target_langs):
    """计算已经翻译的单元格数量"""
    already_translated = 0
    total_cells = 0
    
    logger.info("开始统计已翻译的单元格...")
    
    for row in tqdm(valid_rows, desc="统计已翻译进度"):
        for col_idx, _ in target_langs:
            total_cells += 1
            cell_value = ws.cell(row=row, column=col_idx).value
            if cell_value and str(cell_value).strip():
                already_translated += 1
    
    percentage = (already_translated / total_cells) * 100 if total_cells > 0 else 0
    logger.info(f"已有 {already_translated}/{total_cells} ({percentage:.2f}%) 个单元格已翻译")
    
    return already_translated

def process_excel_with_threading(excel_file=None, output_file=None, source_lang="English", 
                               target_languages=None, api_key_param=None, reference_file=None, 
                               reference_lang=None, reference_column=None):
    """使用多线程处理Excel文件"""
    global translated_count, translation_cancelled, api_key, total_tasks, last_progress_report
    
    # 重置所有计数器和状态
    translated_count = 0
    last_progress_report = 0
    translation_cancelled = False
    api_key = api_key_param
    total_tasks = 0

    try:
        # 读取Excel文件
        wb = load_workbook(excel_file)
        ws = wb.active
        
        # 获取标题行
        header_row = [cell.value for cell in ws[1]]
        max_row = ws.max_row
        
        # 创建新的工作簿
        new_wb = load_workbook(excel_file)
        new_ws = new_wb.active
        
        # 清除现有的目标语言列（如果存在）
        existing_cols = []
        for col_idx, header in enumerate(header_row, 1):
            if header in target_languages:
                existing_cols.append(col_idx)
        
        # 从后往前删除列（避免索引变化）
        for col_idx in sorted(existing_cols, reverse=True):
            new_ws.delete_cols(col_idx)
        
        # 重新获取最后一列的索引
        last_col = new_ws.max_column
        
        # 添加目标语言列
        target_langs = []
        for i, lang in enumerate(target_languages, 1):
            col_idx = last_col + i
            new_ws.cell(row=1, column=col_idx, value=lang)
            target_langs.append((col_idx, lang))
            
        # 获取源语言列索引
        source_col = None
        for col_idx, header in enumerate(header_row, 1):
            # 检查中文名称和英文代码
            if header == source_lang or header == SUPPORTED_LANGUAGES.get(source_lang) or \
               header == [k for k, v in SUPPORTED_LANGUAGES.items() if v == source_lang][0]:
                source_col = col_idx
                break
                
        if source_col is None:
            raise ValueError(f"未找到源语言列: {source_lang}")
            
        # 获取有效行（源语言列有值的行）
        valid_rows = []
        for row in range(2, max_row + 1):
            source_text = ws.cell(row=row, column=source_col).value
            if source_text and str(source_text).strip():
                valid_rows.append((row, str(source_text).strip()))
        
        # 计算总任务数
        total_tasks = len(valid_rows) * len(target_langs)
        
        # 准备参考源数据
        reference_data = {}
        if reference_file and reference_lang:
            ref_wb = load_workbook(reference_file)
            ref_ws = ref_wb.active
            ref_header_row = [cell.value for cell in ref_ws[1]]
            
            # 在参考文件中查找参考语言列
            ref_target_col = None
            for col_idx, header in enumerate(ref_header_row, 1):
                # 检查中文名称
                if header == reference_lang:
                    ref_target_col = col_idx
                    break
                # 检查英文代码
                if header == SUPPORTED_LANGUAGES.get(reference_lang):
                    ref_target_col = col_idx
                    break
                # 如果传入的是英文代码，检查对应的中文名称
                for zh_name, en_code in SUPPORTED_LANGUAGES.items():
                    if en_code == reference_lang and header == zh_name:
                        ref_target_col = col_idx
                        break
                if ref_target_col:
                    break
            
            if ref_target_col is None:
                raise ValueError(f"在参考文件中未找到{reference_lang}列（支持中英文列名）")
            
            # 读取参考数据
            for row in range(2, ref_ws.max_row + 1):
                source = ref_ws.cell(row=row, column=source_col).value
                ref = ref_ws.cell(row=row, column=ref_target_col).value
                if source and ref:
                    reference_data[str(source).strip()] = str(ref).strip()

        # 开始翻译处理
        with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
            for start_idx in range(0, len(valid_rows), batch_size):
                if translation_cancelled:
                    return False

                end_idx = min(start_idx + batch_size, len(valid_rows))
                batch_rows = valid_rows[start_idx:end_idx]
                
                futures = []
                for col_idx, lang in target_langs:
                    texts_to_translate = []
                    rows_to_update = []
                    
                    for row_idx, text in batch_rows:
                        current_text = new_ws.cell(row=row_idx, column=col_idx).value
                        if not current_text or not str(current_text).strip():
                            if reference_file and reference_lang and text in reference_data:
                                texts_to_translate.append((text, reference_data[text]))
                            else:
                                texts_to_translate.append((text, None))
                            rows_to_update.append(row_idx)
                    
                    if texts_to_translate:
                        if reference_file and reference_lang:
                            future = executor.submit(
                                translate_batch_with_reference,
                                texts_to_translate,
                                lang,
                                reference_lang
                            )
                        else:
                            future = executor.submit(
                                translate_batch,
                                [text[0] for text in texts_to_translate],
                                source_lang,
                                lang
                            )
                        futures.append((future, rows_to_update, col_idx))
                
                # 处理翻译结果
                for future, rows, col_idx in futures:
                    if translation_cancelled:
                        return False

                    try:
                        translations = future.result()
                        with excel_lock:
                            for i, row in enumerate(rows):
                                if i < len(translations):
                                    new_ws.cell(row=row, column=col_idx).value = translations[i][1]
                                    translated_count += 1
                                    update_progress_status(translated_count, total_tasks)
                            
                            if translated_count % save_interval == 0:
                                new_wb.save(output_file)
                    except Exception as e:
                        logger.error(f"处理翻译结果时出错: {e}")
        
        # 保存最终结果
        new_wb.save(output_file)
        
        # 更新最终进度
        if not translation_cancelled:
            update_progress_status(total_tasks, total_tasks, True)
        
        return True
        
    except Exception as e:
        logger.error(f"处理Excel文件出错: {e}")
        return False

def translate_batch(texts, source_lang, target_lang, retry_count=0):
    """不使用参考源的批量翻译"""
    if not texts or translation_cancelled:
        return []
    
    try:
        if translation_cancelled:
            return [("[已取消]", "[已取消]") for _ in texts]

        if not api_key:
            raise ValueError("API Key未设置")
            
        client = OpenAI(api_key=api_key, base_url=DEEPSEEK_BASE_URL)
        
        # 构建批量翻译请求
        batch_prompts = []
        for i, text in enumerate(texts):
            batch_prompts.append(f"{i+1}. {source_lang}原文: {text}")
        
        batch_text = "\n\n".join(batch_prompts)
        
        prompt = f"""请将以下{len(texts)}条{source_lang}文本翻译成{target_lang}。
请确保翻译准确、自然、符合目标语言的表达习惯。
仅返回翻译结果，格式为"1. [翻译结果1]"，"2. [翻译结果2]"等，不要有额外解释。

{batch_text}"""

        if translation_cancelled:
            return [("[已取消]", "[已取消]") for _ in texts]

        try:
            response = client.chat.completions.create(
                model="deepseek-chat",
                messages=[
                    {"role": "system", "content": "你是一个精通多语言翻译的专家。只翻译文本，不添加任何解释或附加文本。返回结果保持简洁。"},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.3,
                max_tokens=2000,
            )
        except Exception as e:
            error_msg = str(e).lower()
            if "401" in error_msg and "invalid" in error_msg and "api key" in error_msg:
                raise ValueError("API Key无效或未授权，请检查API Key是否正确")
            raise
        
        # 解析返回结果
        result = response.choices[0].message.content.strip()
        translated_texts = []
        
        # 处理返回的格式，试图匹配"数字. 翻译内容"格式
        lines = result.split("\n")
        for i, text in enumerate(texts):
            found = False
            for line in lines:
                line = line.strip()
                # 尝试匹配格式为 "i+1. 翻译内容"
                if line.startswith(f"{i+1}. ") or line.startswith(f"{i+1}."):
                    translated = line[line.find(".")+1:].strip()
                    translated_texts.append((text, translated))
                    found = True
                    break
            
            if not found:
                # 如果无法匹配格式，则尝试按顺序取出翻译结果
                translated_texts.append((text, "[格式错误]"))
        
        # 确保返回结果数量匹配
        while len(translated_texts) < len(texts):
            translated_texts.append(("[翻译缺失]", "[翻译缺失]"))
            
        return translated_texts
        
    except ValueError as e:
        # API Key相关错误直接向上抛出
        if "api key无效" in str(e).lower():
            raise
        # 其他错误进行重试
        if retry_count < max_retries:
            time.sleep(1)
            return translate_batch(texts, source_lang, target_lang, retry_count + 1)
        return [("[翻译错误]", "[翻译错误]") for _ in texts]
    except Exception as e:
        # 其他错误进行重试
        if retry_count < max_retries:
            time.sleep(1)
            return translate_batch(texts, source_lang, target_lang, retry_count + 1)
        return [("[翻译错误]", "[翻译错误]") for _ in texts]

def set_config(config):
    """设置全局配置参数"""
    global max_workers, batch_size, max_retries, save_interval, progress_interval
    
    max_workers = config.get('max_workers', DEFAULT_MAX_WORKERS)
    batch_size = config.get('batch_size', DEFAULT_BATCH_SIZE)
    max_retries = config.get('max_retries', DEFAULT_MAX_RETRIES)
    save_interval = config.get('save_interval', DEFAULT_SAVE_INTERVAL)
    progress_interval = config.get('progress_interval', DEFAULT_PROGRESS_INTERVAL)

def main():
    logger.info("开始多线程批量翻译Excel文件")
    logger.info(f"线程数: {max_workers}, 批处理大小: {batch_size}, 保存间隔: {save_interval}, 进度显示间隔: {progress_interval}")
    
    start_time = time.time()
    success = process_excel_with_threading()
    end_time = time.time()
    
    if success:
        logger.info(f"Excel翻译成功完成! 总耗时: {end_time - start_time:.2f} 秒")
    else:
        logger.error("Excel翻译过程中出现错误")

if __name__ == "__main__":
    main()